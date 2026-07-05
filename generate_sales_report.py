#!/usr/bin/env python3
"""
Sales/refund report generator from the Udemy statement CSV.

Usage:
    python generate_sales_report.py [path_to_csv]

If no path is given, looks for a single .csv file in the current
directory. Generates a .xlsx with 4 sheets: Marketplace, Udemy
Business, Personal Plan and Total Geral, sorted by course ID.
"""

import sys
import csv
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Course ID -> Name mapping (current names + known aliases/old names).
# Course names are kept in Portuguese on purpose (they match the CSV).
# ---------------------------------------------------------------------------

COURSES = [
    ("1754774", "ASP.NET 2026 do 0 à Azure e GCP com ASP.NET 10 Docker e K8s", []),
    ("1787636", "Organize Suas Finanças Pessoais com Excel Passo a Passo", []),
    ("1860476", "Kindle Desmistificado: Publique seu livro na Amazon", []),
    ("1888598", "Trello 2026: Gestão Otimizada de Equipes e Projetos Pessoais",
     ["Trello 2025: Gestão Otimizada de Equipes e Projetos Pessoais"]),
    ("1921406", "Agile desmistificado com Scrum, XP, Kanban, Spotify e Trello", []),
    ("2113364", "Squad e Spotify Engineering Culture Desmistificado", []),
    ("2178262", "Spring Boot 2026 REST API's do 0 à AWS e GCP c Java e Docker",
     ["Spring Boot 2025 REST API's do 0 à AWS e GCP c Java e Docker"]),
    ("2340578", "Transformação Ágil: Entregue mais e mais Rápido com Scrum", []),
    ("2414176", "Docker e Kubernetes 2026: do Zero à Inteligência Artificial",
     ["Docker do 0 à Maestria: Contêineres Desmistificados com K8s"]),
    ("2453556", "Docker e Kubernetes do 0 à AWS, Azure e GCP c Github Actions", []),
    ("2657590", "REST API's RESTFul from 0 to AWS with Spring Boot and Docker", []),
    ("2988920", "Docker to Amazon AWS Deploy Java & .NET Apps with Travis CI", []),
    ("3606870", "Career Hacking: Atalhos para o sucesso em TI", []),
    ("3881188", "Microsserviços 2026 c. Spring Cloud Boot Kubernetes e Docker",
     ["Microsserviços 2025 c. Spring Cloud Boot Kubernetes e Docker"]),
    ("3881196", "Arquitetura de Microsserviços do 0 com ASP.NET, .NET 6 e C#", []),
    ("4239058", "React JS consumindo REST API RESTful em Spring Boot Java 21", []),
    ("4239060", "React JS consumindo REST API RESTful em ASP.NET 8 e .NET 8", []),
    ("4489848", "REST API's RESTFul do 0 à AWS c. Spring Boot Kotlin e Docker", []),
    ("4495714", "React JS consumindo REST API RESTful em Spring Boot e Kotlin", []),
    ("4564418", "Kotlin para DEVs Java: Aprenda a Linguagem Padrão do Android", []),
    ("4651570", "Agile e Kanban para Times em Home Office com Trello", []),
    ("4651582", "Microsserviços do 0 com Spring Cloud, Kotlin e Docker", []),
    ("4737956", "Agile e Scrum para Times em Home Office com Trello", []),
    ("5026134", "Java Unit Testing com Spring Boot, TDD, Junit e Mockito",
     ["Java Unit Testing com Spring Boot 3, TDD, Junit 5 e Mockito"]),
    ("5385596", "Java Continuous Integration-Delivery c. AWS e Github Actions", []),
    ("5391204", "Java CI e CD com Testes, Microsoft Azure e Github Actions", []),
    ("5391542", "Continuous Deployment c Java GCP Kubernetes e Github Actions", []),
    ("6404313", "Relatórios Profissionais c Java, Spring Boot e JasperReports", []),
    ("6483467", "Spring AI com Spring Boot, Ollama, DeepSeek, MCP e ChatGPT", []),
    ("6519957", "Spring AI c Kotlin Spring Boot ChatGPT MCP Claude e DeepSeek", []),
    ("6519963", "Inteligência Artificial c .NET AI DeepSeek OpenAI e ChatGPT", []),
]


def normalize(name):
    """Strip whitespace and lowercase, so minor formatting differences
    (e.g. 'ASP .NET' vs 'ASP.NET') don't create false mismatches."""
    return "".join(name.lower().split())


# normalized_name -> (id, canonical_name)
NAME_LOOKUP = {}
for cid, canonical, aliases in COURSES:
    NAME_LOOKUP[normalize(canonical)] = (cid, canonical)
    for alias in aliases:
        NAME_LOOKUP[normalize(alias)] = (cid, canonical)


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

SECTION_MARKERS = ["Sales", "Refunds", "Udemy Business", "Personal Plan", "Others"]


def find_csv_path(arg_path):
    if arg_path:
        return arg_path
    candidates = glob.glob("*.csv")
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) == 0:
        sys.exit("No .csv file found in the current directory. "
                  "Pass the path explicitly: python generate_sales_report.py path.csv")
    sys.exit(f"More than one .csv found ({', '.join(candidates)}). "
              f"Specify which one to use: python generate_sales_report.py path.csv")


def find_section_indices(lines):
    idxs = {}
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped in SECTION_MARKERS and stripped not in idxs:
            idxs[stripped] = i
    return idxs


def parse_block(lines, start_idx, end_idx):
    """start_idx = index of the header row (first line after the marker)."""
    block_text = "\n".join(lines[start_idx:end_idx])
    rows = list(csv.reader(block_text.splitlines()))
    if not rows:
        return []
    return rows[1:]  # skip header


def resolve_course(raw_name, unmapped_seen):
    key = normalize(raw_name)
    if key in NAME_LOOKUP:
        return NAME_LOOKUP[key]
    unmapped_seen.add(raw_name)
    return (None, raw_name)


def parse_statement(csv_path):
    with open(csv_path, encoding="utf-8-sig") as f:
        lines = f.read().split("\n")

    idxs = find_section_indices(lines)
    ordered_markers = [m for m in SECTION_MARKERS if m in idxs]

    def section_range(marker):
        start = idxs[marker] + 1
        pos = ordered_markers.index(marker)
        end = idxs[ordered_markers[pos + 1]] if pos + 1 < len(ordered_markers) else len(lines)
        return start, end

    unmapped = set()
    sold, refund, ub, pp = {}, {}, {}, {}
    others_total = 0.0

    if "Sales" in idxs:
        start, end = section_range("Sales")
        for row in parse_block(lines, start, end):
            if len(row) < 13:
                continue
            cid, cname = resolve_course(row[3].strip(), unmapped)
            key = cid or cname
            try:
                sold[key] = sold.get(key, 0.0) + float(row[12])
            except ValueError:
                continue

    if "Refunds" in idxs:
        start, end = section_range("Refunds")
        for row in parse_block(lines, start, end):
            if len(row) < 5:
                continue
            cid, cname = resolve_course(row[2].strip(), unmapped)
            key = cid or cname
            try:
                refund[key] = refund.get(key, 0.0) + abs(float(row[4]))
            except ValueError:
                continue

    ub_available = "Udemy Business" in idxs
    if ub_available:
        start, end = section_range("Udemy Business")
        for row in parse_block(lines, start, end):
            if len(row) < 3:
                continue
            cid, cname = resolve_course(row[0].strip(), unmapped)
            key = cid or cname
            try:
                ub[key] = ub.get(key, 0.0) + float(row[2])
            except ValueError:
                continue

    pp_available = "Personal Plan" in idxs
    if pp_available:
        start, end = section_range("Personal Plan")
        for row in parse_block(lines, start, end):
            if len(row) < 3:
                continue
            cid, cname = resolve_course(row[0].strip(), unmapped)
            key = cid or cname
            try:
                pp[key] = pp.get(key, 0.0) + float(row[2])
            except ValueError:
                continue

    if "Others" in idxs:
        start, end = section_range("Others")
        for row in parse_block(lines, start, end):
            if len(row) < 2:
                continue
            try:
                others_total += float(row[1])
            except ValueError:
                continue

    if unmapped:
        print("WARNING: the following course names in the CSV did not match the "
              "reference list and were kept with a blank ID (please review manually):")
        for name in sorted(unmapped):
            print(f"  - {name}")

    if not ub_available:
        print("WARNING: 'Udemy Business' section not found in the CSV (data not yet "
              "available for this period) - values treated as zero.")
    if not pp_available:
        print("WARNING: 'Personal Plan' section not found in the CSV (data not yet "
              "available for this period) - values treated as zero.")

    return {
        "sold": sold, "refund": refund, "ub": ub, "pp": pp,
        "others_total": others_total,
        "ub_available": ub_available, "pp_available": pp_available,
        "unmapped": unmapped,
    }


def build_rows(parsed):
    known_ids = {cid for cid, _, _ in COURSES}
    keys = set(parsed["sold"]) | set(parsed["refund"]) | set(parsed["ub"]) | set(parsed["pp"]) | known_ids

    rows = []
    for cid, canonical, _ in COURSES:
        s = parsed["sold"].get(cid, 0.0)
        r = parsed["refund"].get(cid, 0.0)
        u = parsed["ub"].get(cid, 0.0)
        p = parsed["pp"].get(cid, 0.0)
        rows.append({"id": cid, "name": canonical, "sold": s, "refund": r, "ub": u, "pp": p})

    # Unmapped course names (unrecognized) go at the end, with a blank ID.
    extra_keys = [k for k in keys if k not in known_ids]
    for key in sorted(extra_keys):
        s = parsed["sold"].get(key, 0.0)
        r = parsed["refund"].get(key, 0.0)
        u = parsed["ub"].get(key, 0.0)
        p = parsed["pp"].get(key, 0.0)
        rows.append({"id": "??", "name": f"{key} (unmapped - please review)", "sold": s, "refund": r, "ub": u, "pp": p})

    return rows


# ---------------------------------------------------------------------------
# XLSX generation.
# Sheet names, headers and labels below are in Portuguese on purpose -
# this is the output formatting the end user reads.
# ---------------------------------------------------------------------------

THIN_BORDER = Border(*(Side(style="thin", color="D0D0D0") for _ in range(4)))
HEADER_FILL = PatternFill("solid", start_color="F8F8F8", end_color="F8F8F8")
ALT_FILL = PatternFill("solid", start_color="FAFAFA", end_color="FAFAFA")
TOTAL_FILL = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")
BOLD = Font(bold=True)
FONT_NAME = "Arial"


def style_money_cell(cell, value):
    cell.number_format = "#,##0.00"
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(name=FONT_NAME, color=("2E7D32" if value > 0 else "D32F2F" if value < 0 else "555555"))
    cell.border = THIN_BORDER


def write_table(ws, headers, data_rows, money_cols, others_row=None):
    """headers: list of column titles. data_rows: list of dicts with 'id', 'name'
    and the numeric columns. money_cols: list of (header_index, dict_key) for
    monetary columns."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT_NAME, bold=True)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right" if j > 2 else "left")

    r = 2
    all_rows = data_rows + ([others_row] if others_row else [])
    for row in all_rows:
        fill = None if (r - 2) % 2 == 0 else ALT_FILL
        id_cell = ws.cell(row=r, column=1, value=row["id"])
        id_cell.font = Font(name=FONT_NAME)
        id_cell.border = THIN_BORDER
        if fill:
            id_cell.fill = fill
        name_cell = ws.cell(row=r, column=2, value=row["name"])
        name_cell.font = Font(name=FONT_NAME)
        name_cell.border = THIN_BORDER
        if fill:
            name_cell.fill = fill
        for col_idx, key in money_cols:
            val = row.get(key, 0.0)
            cell = ws.cell(row=r, column=col_idx, value=round(val, 2))
            style_money_cell(cell, val)
            if fill:
                cell.fill = fill
        r += 1

    total_row = r
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=2).border = THIN_BORDER
    for col_idx, _ in money_cols:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col_idx, value=formula)
        cell.number_format = "#,##0.00"
        cell.font = BOLD
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 60 if col_idx == 2 else 16


def generate_xlsx(parsed, out_path):
    rows = build_rows(parsed)
    others_total = round(parsed["others_total"], 2)

    wb = Workbook()

    # --- Marketplace ---
    ws1 = wb.active
    ws1.title = "Marketplace"
    mp_rows = [{**r, "net": round(r["sold"] - r["refund"], 2)} for r in rows]
    others_mp = {"id": "-", "name": "Others (Subscription Bonuses)", "sold": others_total, "refund": 0.0, "net": others_total}
    write_table(
        ws1,
        ["ID", "Curso", "Total Vendido", "Reembolsos", "Total sem Reembolsos"],
        mp_rows,
        [(3, "sold"), (4, "refund"), (5, "net")],
        others_row=others_mp,
    )

    # --- Udemy Business ---
    ws2 = wb.create_sheet("Udemy Business")
    if parsed["ub_available"]:
        write_table(ws2, ["ID", "Curso", "Instructor Share"], rows, [(3, "ub")])
    else:
        ws2.cell(row=1, column=1, value="Seção 'Udemy Business' ainda não disponível neste CSV.").font = Font(name=FONT_NAME, italic=True)

    # --- Personal Plan ---
    ws3 = wb.create_sheet("Personal Plan")
    if parsed["pp_available"]:
        write_table(ws3, ["ID", "Curso", "Instructor Share"], rows, [(3, "pp")])
    else:
        ws3.cell(row=1, column=1, value="Seção 'Personal Plan' ainda não disponível neste CSV.").font = Font(name=FONT_NAME, italic=True)

    # --- Total Geral ---
    ws4 = wb.create_sheet("Total Geral")
    tg_rows = [{**r, "net": round(r["sold"] - r["refund"], 2),
                "total": round(r["sold"] - r["refund"] + r["ub"] + r["pp"], 2)} for r in rows]
    others_tg = {"id": "-", "name": "Others (Subscription Bonuses)", "net": others_total, "ub": 0.0, "pp": 0.0, "total": others_total}
    write_table(
        ws4,
        ["ID", "Curso", "Marketplace", "Udemy Business", "Personal Plan", "Total"],
        tg_rows,
        [(3, "net"), (4, "ub"), (5, "pp"), (6, "total")],
        others_row=others_tg,
    )

    wb.save(out_path)


def main():
    arg_path = sys.argv[1] if len(sys.argv) > 1 else None
    csv_path = find_csv_path(arg_path)
    parsed = parse_statement(csv_path)
    out_path = os.path.splitext(os.path.basename(csv_path))[0] + "_report.xlsx"
    generate_xlsx(parsed, out_path)
    print(f"Report generated: {out_path}")


if __name__ == "__main__":
    main()
